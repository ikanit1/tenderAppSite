# -*- coding: utf-8 -*-
"""Parse Akuvox Smart Home xlsx: extract products, images, convert USD->KZT, russify. Output JSON for frontend."""
import json
import re
import zipfile
import openpyxl
from pathlib import Path
from io import BytesIO

try:
    from openpyxl.reader.drawings import find_images
except ImportError:
    find_images = None

USD_TO_KZT = 500  # курс для конвертации
XLSX_PATH = Path(r"f:\downloads\Telegram Desktop\Akuvox Smart Home Product MSRP Price List update0831-.xlsx")
OUT_PATH = Path(__file__).resolve().parent.parent / "src" / "shared" / "content" / "akuvoxSmartSystems.json"
OUT_IMAGES_DIR = Path(__file__).resolve().parent.parent / "public" / "akuvox"

# Колонки, в которых может быть картинка (0-based: A=0 .. J=9), чтобы не терять картинки в D,E,F...
PICTURE_COLS = tuple(range(10))
# Первая строка данных (0-based)
DATA_START_ROW = 4

# Расширенный словарь русификации (все ключевые термины из документа)
DESC_RU = {
    "All In One Panel": "Сенсорная панель All-in-One",
    "Touch Screen": "сенсорный экран",
    "2-gang": "2-клавишный",
    "1-gang": "1-клавишный",
    "POE": "PoE питание",
    "Fancoil": "контроль фанкойла",
    "0-10V": "управление 0–10 В",
    "24V": "24 В",
    "Dimmer": "диммер",
    "heating": "обогрев",
    "Keypads": "клавиатура",
    "Desktop stand": "настольная подставка",
    "Center Panel": "Центральная панель",
    "Linux based": "на базе Linux",
    "Android based": "на базе Android",
    "2.1-inch": "экран 2.1\"",
    "touch screen": "сенсорный экран",
    "Excellent performance": "Отличная производительность",
    "Android 10 system": "система Android 10",
    "Android 12 system": "система Android 12",
    "Android 13 system": "система Android 13",
    "energy saving mode": "режим энергосбережения",
    "Smart intercom": "Умный домофон",
    "ZigBee": "ZigBee",
    "gate": "шлюз",
    "supported": "поддержка",
    "support": "поддержка",
    "standard": "стандарт",
    "EU standard": "стандарт EU",
    "US standard": "стандарт US",
    "suitable for the whole family": "удобно для всей семьи",
    "touch buttons": "сенсорные кнопки",
    "switch": "выключатель",
    "replacer": "замена",
    "thermostat": "термостат",
    "thermostat control": "управление термостатом",
    "KNX": "KNX",
    "Controller": "контроллер",
    "control": "управление",
    "version": "версия",
    "PCS": "шт.",
    "Box Quantity": "Количество в коробке",
    "Model Name": "Модель",
    "Description": "Описание",
    "MSRP": "Рекомендуемая цена",
    "Notes": "Примечания",
    "quotation": "прайс",
    "based on": "на основе",
    "Warehouse": "Склад",
    "innovation": "инновации",
    "technology": "технология",
    "replace conventional": "замена обычного",
    "conventional": "обычный",
    "flush mounted": "встраиваемый",
    "flush mounted box": "коробка для встраивания",
    "built-in": "встроенный",
    "built in": "встроенный",
    "sensors": "датчики",
    "sensor": "датчик",
    "multi-": "мульти-",
    "multi": "мульти",
    "ALL IN ONE": "всё в одном",
    "all in one": "всё в одном",
    "Kinds of sensors built-in": "Встроены различные датчики",
    "home automation": "умный дом",
    "digital input": "цифровой вход",
    "input port": "входной порт",
    "port": "порт",
    "way": "путь",
    "gateway": "шлюз",
    "replacer": "замена",
    "replace": "заменить",
    "function": "функция",
    "excellent": "отличный",
    "performance": "производительность",
    "with": "с",
    "and": "и",
    "for": "для",
    "Kinds of": "Различные",
    "mounted": "монтаж",
    "box": "коробка",
    "RS485": "RS485",
    "FAN COIL": "фанкойл",
    "FAN": "вентилятор",
    "COIL": "теплообменник",
    "supported": "поддерживается",
    "flush": "встраиваемый",
    "mounted box": "монтажная коробка",
}


def desc_to_ru(text):
    """Полная русификация описания: замена всех фраз из словаря по всему тексту."""
    if not text or not isinstance(text, str):
        return ""
    t = text
    # Сортируем по длине фразы (длинные первыми), чтобы не затирать короткие вхождения внутри длинных
    for en, ru in sorted(DESC_RU.items(), key=lambda x: -len(x[0])):
        t = re.sub(re.escape(en), ru, t, flags=re.I)
    return t.strip()


def model_clean(s):
    if not s:
        return ""
    s = str(s).strip()
    return s.split("\n")[0].strip() if "\n" in s else s


def sanitize_filename(model):
    """Безопасное имя файла для модели."""
    s = re.sub(r"[^\w\-]", "_", model)
    return s[:80].strip("_") or "product"


def category_from_model(model):
    m = model.upper()
    if m.startswith("DS01"):
        return "Аксессуары"
    if m.startswith("PS51"):
        return "Сенсорные панели PS51"
    if m.startswith("PS52"):
        return "Сенсорные панели PS52"
    if m.startswith("KS53"):
        return "Панели с клавиатурой KS53"
    if m.startswith("KS41"):
        return "Центральные панели KS41"
    if m.startswith("RT61"):
        return "Компактные панели RT61"
    return "Умные панели"


def get_image_extension(img):
    """Определить расширение по типу изображения в openpyxl."""
    try:
        if hasattr(img, "ref") and img.ref:
            p = getattr(img, "path", None) or getattr(img, "_path", None)
            if p and isinstance(p, (str, Path)):
                return Path(p).suffix.lower() or ".png"
    except Exception:
        pass
    try:
        if hasattr(img, "_data"):
            data = img._data()
            if data[:8] == b"\x89PNG\r\n\x1a\n":
                return ".png"
            if data[:2] == b"\xff\xd8":
                return ".jpg"
            if data[:6] in (b"GIF87a", b"GIF89a"):
                return ".gif"
    except Exception:
        pass
    return ".png"


def _get_anchor_row_col(anchor):
    """Из anchor (OneCellAnchor/TwoCellAnchor) получить (row, col). Для AbsoluteAnchor — None."""
    if anchor is None:
        return None
    from_ = getattr(anchor, "_from", None) or getattr(anchor, "from", None)
    if from_ is None:
        return None
    row = getattr(from_, "row", 0)
    col = getattr(from_, "col", 0)
    return (int(row), int(col))


def _get_first_sheet_drawing_path(archive):
    """По архиву xlsx получить путь к drawing первого листа (xl/drawings/drawingN.xml)."""
    try:
        if "xl/workbook.xml" not in archive.namelist():
            return None
        from xml.etree import ElementTree as ET
        wb_xml = archive.read("xl/workbook.xml")
        root = ET.fromstring(wb_xml)
        ns = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main", "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships"}
        sheets = root.findall(".//main:sheet", ns) or root.findall(".//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheet")
        if not sheets:
            return None
        first = sheets[0]
        rId = first.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id") or first.get("r:id")
        if not rId:
            return None
        rels_path = "xl/_rels/workbook.xml.rels"
        if rels_path not in archive.namelist():
            return None
        rels_xml = archive.read(rels_path)
        rels = ET.fromstring(rels_xml)
        for rel in rels:
            rid = rel.get("Id")
            if rid == rId:
                target = rel.get("Target", "")
                if "worksheets" in target:
                    sheet_path = target if target.startswith("xl/") else "xl/" + target.lstrip("/")
                    sheet_rels_path = sheet_path.replace("xl/worksheets/", "xl/worksheets/_rels/") + ".rels"
                    if sheet_rels_path not in archive.namelist():
                        return None
                    sheet_rels = ET.fromstring(archive.read(sheet_rels_path))
                    for srel in sheet_rels:
                        stype = srel.get("Type", "")
                        if "drawing" in stype:
                            draw_target = srel.get("Target", "").replace("\\", "/")
                            parts = sheet_rels_path.replace("\\", "/").split("/")[:-1] + draw_target.split("/")
                            resolved = []
                            for p in parts:
                                if p == "..":
                                    if resolved:
                                        resolved.pop()
                                else:
                                    resolved.append(p)
                            full = "/".join(resolved)
                            if full in archive.namelist():
                                return full
                            alt = "xl/drawings/" + draw_target.split("/")[-1]
                            if alt in archive.namelist():
                                return alt
                    return None
                break
    except Exception:
        pass
    return None


def _images_from_find_images(archive, drawing_path):
    """Из drawing XML через find_images получить список Image (с anchor)."""
    if not find_images or not drawing_path or drawing_path not in archive.namelist():
        return []
    try:
        _charts, images = find_images(archive, drawing_path)
        return list(images) if images else []
    except Exception:
        return []


def _images_from_all_drawings(archive):
    """Обойти все xl/drawings/drawing*.xml и собрать изображения (для листа без явной привязки — по порядку)."""
    out = []
    for name in sorted(archive.namelist()):
        if name.startswith("xl/drawings/") and name.endswith(".xml") and "drawing" in name:
            if find_images:
                try:
                    _c, imgs = find_images(archive, name)
                    out.extend(imgs or [])
                except Exception:
                    pass
    return out


def _media_files_from_archive(archive):
    """Все файлы из xl/media/ — возвращает список (bytes, extension), отсортированный по имени (порядок как в книге)."""
    result = []
    for name in sorted(archive.namelist()):
        if not name.startswith("xl/media/"):
            continue
        if name.count("/") != 2:
            continue
        try:
            data = archive.read(name)
            ext = Path(name).suffix.lower() or ".png"
            if ext in (".emf", ".wmf"):
                continue
            if ext not in (".png", ".jpg", ".jpeg", ".gif", ".bmp", ".tiff"):
                ext = ".png"
            result.append((data, ext if ext.startswith(".") else "." + ext))
        except Exception:
            pass
    return result


def build_row_to_image_map(ws, archive=None):
    """Строит маппинг: номер строки (0-based) -> объект Image.
    Источники: 1) ws._images, 2) find_images(drawing) по первому листу, 3) все drawing*.xml.
    Учитывает OneCellAnchor, TwoCellAnchor; колонки A,B,C; AbsoluteAnchor — в список без привязки.
    Возвращает (row_to_image, absolute_anchored, media_fallback_list).
    media_fallback_list — список (bytes, ext) из xl/media/ для подстановки по порядку, если передан archive.
    """
    row_to_image = {}
    absolute_anchored = []
    candidates = []

    def add_image(img):
        try:
            anchor = getattr(img, "anchor", None)
            if anchor is None:
                absolute_anchored.append(img)
                return
            rc = _get_anchor_row_col(anchor)
            if rc is None:
                absolute_anchored.append(img)
                return
            row, col = rc
            if col not in PICTURE_COLS:
                absolute_anchored.append(img)
                return
            candidates.append((row, col, img))
        except Exception:
            pass

    for img in getattr(ws, "_images", []):
        add_image(img)

    media_fallback_list = []
    if archive is not None:
        drawing_path = _get_first_sheet_drawing_path(archive)
        for img in _images_from_find_images(archive, drawing_path):
            add_image(img)
        if not drawing_path:
            for img in _images_from_all_drawings(archive):
                add_image(img)
        media_fallback_list = _media_files_from_archive(archive)

    # Приоритет колонки: B(1) > A(0) > C(2) > остальные
    col_priority = {1: 0, 0: 1, 2: 2}
    for row, col, img in candidates:
        existing = row_to_image.get(row)
        if existing is None:
            row_to_image[row] = img
        else:
            existing_rc = _get_anchor_row_col(getattr(existing, "anchor", None))
            cur_prio = col_priority.get(col, 99)
            exist_prio = col_priority.get(existing_rc[1], 99) if existing_rc else 99
            if cur_prio < exist_prio:
                absolute_anchored.append(existing)
                row_to_image[row] = img
            else:
                absolute_anchored.append(img)
    return row_to_image, absolute_anchored, media_fallback_list


class _BlobImage:
    """Минимальная обёртка над (bytes, ext) для подстановки картинки из xl/media/."""
    def __init__(self, data, ext=".png"):
        self._blob = data
        self._ext = ext if ext.startswith(".") else "." + ext
    def _data(self):
        return self._blob
    def _extension(self):
        return self._ext


def main():
    # Загрузка без read_only, чтобы получить доступ к ws._images; data_only для значений ячеек
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))

    archive = None
    if XLSX_PATH.exists():
        try:
            archive = zipfile.ZipFile(XLSX_PATH, "r")
        except Exception:
            pass
    row_to_image, absolute_anchored, media_fallback_list = build_row_to_image_map(ws, archive)
    if archive:
        try:
            archive.close()
        except Exception:
            pass
    total_imgs_in_sheet = len(getattr(ws, "_images", []))
    print(f"Images: ws._images={total_imgs_in_sheet}, mapped rows={len(row_to_image)}, absolute={len(absolute_anchored)}, media fallback={len(media_fallback_list)}")

    OUT_IMAGES_DIR.mkdir(parents=True, exist_ok=True)
    used_filenames = set()

    products = []
    for i, row in enumerate(rows):
        if i < DATA_START_ROW:
            continue
        model_raw = row[0] if len(row) > 0 else None
        desc_en = row[2] if len(row) > 2 else None
        msrp = row[3] if len(row) > 3 else None
        if not model_raw or not str(model_raw).strip():
            continue
        try:
            price_usd = float(msrp) if msrp is not None else 0
        except (TypeError, ValueError):
            price_usd = 0
        if price_usd <= 0:
            continue
        model = model_clean(model_raw)
        category = category_from_model(model)
        image_path = ""
        # Строка может быть 0- или 1-базовая в Excel — пробуем i, i+1, i-1
        img = row_to_image.get(i) or row_to_image.get(i + 1) or (row_to_image.get(i - 1) if i > 0 else None)
        if img is None and absolute_anchored:
            img = absolute_anchored.pop(0)
        if img is None and media_fallback_list:
            blob, ext = media_fallback_list.pop(0)
            img = _BlobImage(blob, ext)
        if img is not None:
            try:
                data = img._data()
                ext = (img._extension() if getattr(img, "_extension", None) else None) or get_image_extension(img)
                base = sanitize_filename(model)
                fname = base + ext
                if fname in used_filenames:
                    fname = f"{base}_{i}{ext}"
                used_filenames.add(fname)
                out_file = OUT_IMAGES_DIR / fname
                with open(out_file, "wb") as f:
                    f.write(data)
                image_path = f"/akuvox/{fname}"
            except Exception:
                pass

        products.append({
            "model": model,
            "descriptionRu": desc_to_ru(desc_en) or (str(desc_en)[:300] if desc_en else ""),
            "priceUsd": round(price_usd, 2),
            "priceKzt": int(round(price_usd * USD_TO_KZT)),
            "category": category,
            "image": image_path,
        })

    wb.close()

    by_category = {}
    for p in products:
        cat = p["category"]
        if cat not in by_category:
            by_category[cat] = []
        by_category[cat].append(p)

    order = [
        "Сенсорные панели PS51",
        "Сенсорные панели PS52",
        "Панели с клавиатурой KS53",
        "Центральные панели KS41",
        "Компактные панели RT61",
        "Аксессуары",
        "Умные панели",
    ]
    categories = [{"id": re.sub(r"\s+", "-", c.lower()), "title": c, "products": by_category.get(c, [])} for c in order if by_category.get(c)]
    for c in by_category:
        if c not in order:
            categories.append({"id": re.sub(r"\s+", "-", c.lower()), "title": c, "products": by_category[c]})

    out = {
        "title": "",
        "subtitle": "",
        "usdToKzt": USD_TO_KZT,
        "categories": categories,
    }
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT_PATH, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)
    with_images = sum(1 for p in products if p.get("image"))
    print(f"Written {len(products)} products, {len(categories)} categories to {OUT_PATH}")
    print(f"Images extracted: {with_images} / {len(products)} (saved to {OUT_IMAGES_DIR})")
    if with_images < len(products) and total_imgs_in_sheet > with_images:
        print(f"Hint: {total_imgs_in_sheet - with_images} image(s) in file did not match data rows (check column/row in Excel).")


if __name__ == "__main__":
    main()
