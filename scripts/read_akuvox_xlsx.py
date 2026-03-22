"""Read Akuvox Smart Home xlsx and output JSON structure (first 25 rows, first 20 cols)."""
import json
import openpyxl

path = r"f:\downloads\Telegram Desktop\Akuvox Smart Home Product MSRP Price List update0831-.xlsx"
wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
out = {"sheet_names": wb.sheetnames}
rows = []
ws = wb[wb.sheetnames[0]]
for i, row in enumerate(ws.iter_rows(max_row=25, values_only=True)):
    # first 20 columns only, truncate long strings
    cells = list(row)[:20]
    rows.append([str(v)[:60] if v is not None else None for v in cells])
out["first_sheet_rows"] = rows
wb.close()
with open("akuvox_preview.json", "w", encoding="utf-8") as f:
    json.dump(out, f, ensure_ascii=False, indent=2)
print("Done. See akuvox_preview.json")
