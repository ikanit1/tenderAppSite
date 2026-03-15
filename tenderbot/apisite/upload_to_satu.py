#!/usr/bin/env python3
"""
Выгрузка товаров на SATU.KZ через Public API.

Единственный источник данных для товаров и изображений — папки portal_export:
  product.json (название, цена, описание, атрибуты, модель), name.txt, description.txt, image_*.jpg.
В API SATU передаются только эти данные; другие источники не используются.

Документация API: https://public-api.docs.satu.kz/ (Products, Orders, OrderStatus и др.)

Переменные окружения (или .env в apisite):
  SATU_API_TOKEN — Bearer-токен доступа к API (получить у сотрудника портала).
  SATU_API_BASE_URL — базовый URL API (по умолчанию https://my.satu.kz/api/v1).

Запуск из папки apisite:
  python upload_to_satu.py                          # сухой прогон (без отправки)
  python upload_to_satu.py --upload                 # POST products/edit (только обновляет существующие по external_id)
  python upload_to_satu.py --import-file            # создать новые товары: файл и POST products/import_file
  python upload_to_satu.py --import-file --limit 50  # импорт первых 50 товаров
  python upload_to_satu.py --export-xlsx file.xlsx   # Excel для ручного импорта (данные из portal_export)
  python upload_to_satu.py --upload-images           # Дозагрузить изображения на уже созданные в SATU товары
  python upload_to_satu.py --full                    # Импорт + названия, описания, цены и фото из portal_export
"""
import json
import logging
import sys
from pathlib import Path

_apisite_dir = Path(__file__).resolve().parent
if str(_apisite_dir) not in sys.path:
    sys.path.insert(0, str(_apisite_dir))

from config import SATU_API_TOKEN
from satu_api import request as satu_request, check_auth as satu_check_auth, product_edit as satu_product_edit, import_file as satu_import_file_api

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
logger = logging.getLogger(__name__)

PORTAL_EXPORT = _apisite_dir / "portal_export"

# Цена в product.json часто в формате "тиын" или tenge*10000
PRICE_DIVISOR = 10000  # 144007700 -> 14400.77


def _price_to_tenge(raw) -> float:
    try:
        v = float(raw)
        if v <= 0:
            return 0.0
        if v >= 1_000_000:
            return round(v / PRICE_DIVISOR, 2)
        return round(v, 2)
    except (TypeError, ValueError):
        return 0.0


def _attributes_to_text(attrs: dict) -> str:
    if not attrs:
        return ""
    return "\n".join(f"{k}: {v}" for k, v in attrs.items())


def load_products_from_portal(limit: int | None = None):
    """Читает все товары из portal_export. Возвращает список dict с полями name, model, description, price_tenge, image_paths."""
    products = []
    if not PORTAL_EXPORT.is_dir():
        logger.error("Папка portal_export не найдена: %s", PORTAL_EXPORT)
        return products

    subdirs = [d for d in PORTAL_EXPORT.iterdir() if d.is_dir()]
    for folder in sorted(subdirs):
        if limit is not None and len(products) >= limit:
            break
        product_json = folder / "product.json"
        if not product_json.exists():
            continue
        try:
            data = json.loads(product_json.read_text(encoding="utf-8"))
        except Exception as e:
            logger.warning("Не удалось прочитать %s: %s", product_json, e)
            continue

        name = (data.get("name") or "").strip() or (folder / "name.txt").read_text(encoding="utf-8").strip() if (folder / "name.txt").exists() else ""
        if not name:
            name = folder.name

        model = (data.get("model") or "").strip() or folder.name
        desc_html = (data.get("description_html") or "").strip()
        desc_plain = (data.get("description") or "").strip()
        desc_file = folder / "description.txt"
        if desc_file.exists():
            desc_plain = desc_file.read_text(encoding="utf-8").strip() or desc_plain
        description = desc_html if desc_html else desc_plain
        attrs = data.get("attributes") or {}
        if attrs:
            description = (description + "\n\n" + _attributes_to_text(attrs)).strip()

        price_tenge = _price_to_tenge(data.get("price"))

        image_names = data.get("images") or []
        image_paths = []
        for im in image_names:
            p = folder / im
            if p.exists():
                image_paths.append(p)

        if not image_paths:
            for f in sorted(folder.iterdir()):
                if f.suffix.lower() in (".jpg", ".jpeg", ".png", ".webp") and f.name.startswith("image"):
                    image_paths.append(f)

        quantity = 1
        for key in ("quantity", "stock", "qty"):
            if data.get(key) is not None:
                try:
                    quantity = max(0, int(data[key]))
                    break
                except (TypeError, ValueError):
                    pass

        products.append({
            "name": name,
            "model": model,
            "description": description,
            "price_tenge": price_tenge,
            "image_paths": image_paths,
            "folder": folder.name,
            "quantity": quantity,
        })
    return products


def upload_products_to_satu(products: list, token: str, dry_run: bool = True, images_only: bool = False):
    """
    Отправляет товары в SATU (products/edit). При dry_run только логирует.
    images_only: если True, отправляем только товары, у которых есть изображения в portal_export.
    """
    if not dry_run and not token:
        logger.error("SATU_API_TOKEN не задан. Укажите в .env или переменной окружения.")
        return 0, 0

    if not dry_run and not satu_check_auth(token):
        logger.error("Проверка токена SATU не пройдена. Проверьте SATU_API_TOKEN и доступность API.")
        return 0, len(products)

    if images_only:
        products = [p for p in products if p.get("image_paths")]
        logger.info("Товаров с изображениями: %s", len(products))

    ok = 0
    err = 0
    for i, p in enumerate(products):
        name, model, description, price_tenge, image_paths = (
            p["name"], p["model"], p["description"], p["price_tenge"], p["image_paths"]
        )
        if dry_run:
            logger.info("[dry-run] Товар %s: %s | %s ₸ | изображений: %s", i + 1, name, price_tenge, len(image_paths))
            ok += 1
            continue

        # Все данные и изображения — только из portal_export (products/edit по external_id)
        try:
            success, resp = satu_product_edit(
                token,
                name=name,
                description=description,
                price=price_tenge,
                external_id=model,
                image_paths=image_paths if image_paths else None,
            )
        except Exception as e:
            logger.warning("Ошибка запроса для %s: %s", model, e)
            err += 1
            continue

        if success:
            ok += 1
            logger.info("Ок: %s (%s) | фото: %s", name, model, len(image_paths))
        else:
            err += 1
            logger.warning("SATU ответ %s для %s: %s", getattr(resp, "status_code", ""), model, (resp.text[:300] if resp else ""))
    return ok, err


def export_products_to_xlsx(products: list, output_path: Path) -> None:
    """Сохраняет товары в Excel для ручного импорта в кабинет SATU (Товары — Импорт)."""
    try:
        import openpyxl
    except ImportError:
        logger.error("Установите openpyxl: pip install openpyxl")
        sys.exit(1)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Товары"
    headers = ["Наименование", "Описание", "Цена (₸)", "Артикул/Модель", "Путь к папке изображений"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    for row_idx, p in enumerate(products, 2):
        ws.cell(row=row_idx, column=1, value=p["name"])
        ws.cell(row=row_idx, column=2, value=(p["description"] or "")[:32000])
        ws.cell(row=row_idx, column=3, value=p["price_tenge"])
        ws.cell(row=row_idx, column=4, value=p["model"])
        ws.cell(row=row_idx, column=5, value=p["folder"])
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    logger.info("Сохранено в %s: %s товаров", output_path, len(products))


def _build_import_xlsx(products: list, limit: int | None = None):
    """Формирует XLSX для импорта SATU: Название позиции, Описание, Цена, Единица измерения, Валюта, Артикул."""
    import openpyxl
    if limit:
        products = products[:limit]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Товары"
    headers = ["Название позиции", "Описание", "Цена", "Единица измерения", "Валюта", "Артикул"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    for row_idx, p in enumerate(products, 2):
        ws.cell(row=row_idx, column=1, value=p["name"])
        ws.cell(row=row_idx, column=2, value=(p["description"] or "")[:32000])
        ws.cell(row=row_idx, column=3, value=p["price_tenge"])
        ws.cell(row=row_idx, column=4, value="шт")
        ws.cell(row=row_idx, column=5, value="KZT")
        ws.cell(row=row_idx, column=6, value=p["model"])
    return wb, len(products)


def _save_import_xlsx(products: list, output_path: Path, limit: int | None = None) -> None:
    """Сохраняет файл импорта SATU на диск для ручной загрузки в кабинете."""
    try:
        wb, count = _build_import_xlsx(products, limit=limit)
    except Exception as e:
        logger.error("Ошибка формирования файла: %s", e)
        raise
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    logger.info("Сохранено %s позиций в %s", count, output_path)


def upload_import_file_to_satu(products: list, token: str, limit: int | None = None) -> bool:
    """
    Формирует файл импорта (XLSX) и отправляет POST /products/import_file.
    Используется для создания новых товаров (в отличие от products/edit).
    """
    try:
        import openpyxl
    except ImportError:
        logger.error("Установите openpyxl: pip install openpyxl")
        return False
    if not token:
        logger.error("SATU_API_TOKEN не задан.")
        return False
    if not satu_check_auth(token):
        return False

    wb, count = _build_import_xlsx(products, limit=limit)
    import_path = _apisite_dir / "data" / "satu_import.xlsx"
    import_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(import_path)
    logger.info("Файл сохранён: %s (%s позиций)", import_path, count)

    try:
        resp = satu_import_file_api(token, import_path, timeout=120)
    except Exception as e:
        logger.warning("Ошибка запроса import_file: %s", e)
        return False

    if resp.status_code in (200, 201, 202):
        logger.info("Импорт отправлен. Ответ: %s %s", resp.status_code, resp.text[:500])
        return True
    logger.warning("import_file ответ: %s %s", resp.status_code, resp.text[:500])
    return False


def full_upload_to_satu(products: list, token: str, limit: int | None = None) -> None:
    """
    Всё в одном: название, описание, цена, изображения — собираются из portal_export и отправляются на SATU.
    1) POST /products/import_file — создаёт позиции (название, описание, цена, ед. изм., валюта, артикул).
    2) Для каждого товара POST /products/edit — обновляет позицию по артикулу и прикрепляет изображения.
    """
    if limit:
        products = products[:limit]
    logger.info("Шаг 1/2: импорт файла (название, описание, цена, артикул)...")
    import_ok = upload_import_file_to_satu(products, token, limit=None)
    if not import_ok:
        logger.warning("Импорт через API не прошёл (возможно, ограничение на одновременные импорты). Продолжаем шаг 2 — обновление с фото по артикулу.")
    else:
        logger.info("Импорт принят. Ждём 5 сек перед дозагрузкой изображений...")
        import time
        time.sleep(5)
    logger.info("Шаг 2/2: отправка названия, описания, цены и изображений по каждому товару...")
    ok, err = upload_products_to_satu(products, token, dry_run=False, images_only=False)
    logger.info("Итого: успешно %s, ошибок %s", ok, err)


def main():
    import argparse
    parser = argparse.ArgumentParser(description="Выгрузка товаров portal_export на SATU.KZ")
    parser.add_argument("--upload", action="store_true", help="Отправить через products/edit (только обновляет существующие)")
    parser.add_argument("--import-file", action="store_true", help="Создать файл и отправить в products/import_file (если SATU не блокирует одновременные импорты)")
    parser.add_argument("--export-import-xlsx", type=Path, default=None, metavar="FILE", help="Только сохранить файл импорта (Название позиции, Цена, Ед. изм., Валюта) для ручной загрузки в кабинете")
    parser.add_argument("--upload-images", action="store_true", help="Дозагрузить изображения на уже созданные в SATU товары")
    parser.add_argument("--full", action="store_true", help="Всё в одном: импорт + названия, описания, цены и изображения на SATU")
    parser.add_argument("--limit", type=int, default=None, help="Максимум товаров для обработки")
    parser.add_argument("--export-xlsx", type=Path, default=None, metavar="FILE", help="Сохранить товары в Excel для ручного импорта в SATU")
    args = parser.parse_args()

    products = load_products_from_portal(limit=args.limit)
    logger.info("Загружено товаров из portal_export: %s", len(products))
    if not products:
        sys.exit(1)

    if args.export_xlsx:
        export_products_to_xlsx(products, args.export_xlsx)
        return

    if args.export_import_xlsx:
        _save_import_xlsx(products, args.export_import_xlsx, args.limit)
        return

    if args.import_file:
        ok = upload_import_file_to_satu(products, SATU_API_TOKEN, limit=args.limit)
        sys.exit(0 if ok else 1)

    if args.upload_images:
        ok, err = upload_products_to_satu(products, SATU_API_TOKEN, dry_run=False, images_only=False)
        logger.info("Итого (изображения): успешно %s, ошибок %s", ok, err)
        return

    if args.full:
        full_upload_to_satu(products, SATU_API_TOKEN, limit=args.limit)
        return

    dry_run = not args.upload
    if dry_run:
        logger.info("Режим dry-run. Для отправки укажите --upload.")

    ok, err = upload_products_to_satu(products, SATU_API_TOKEN, dry_run=dry_run)
    logger.info("Итого: успешно %s, ошибок %s", ok, err)


if __name__ == "__main__":
    main()
