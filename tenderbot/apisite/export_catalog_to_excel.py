#!/usr/bin/env python3
"""
Скрипт экспорта каталога товаров из API Complex в Excel.

Товары берутся из того же источника, что и каталог (B2B API complex.com.kz).
Запуск из папки apisite: python export_catalog_to_excel.py

Результат: data/catalog_export.xlsx (или путь из аргумента --output).
"""
import json
import logging
import sys
from pathlib import Path

# Чтобы при запуске из корня проекта работали импорты из apisite
_apisite_dir = Path(__file__).resolve().parent
if str(_apisite_dir) not in sys.path:
    sys.path.insert(0, str(_apisite_dir))

from config import B2B_API_URL, API_KEY, CACHE_FILE
from b2b_client import B2BClient
from price_manager import calculate_final_price

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
logger = logging.getLogger(__name__)

# Путь к items.json для категорий (как в main.py)
PORTAL_ITEMS_JSON = _apisite_dir / "portal_export" / "items.json"
DEFAULT_OUTPUT = _apisite_dir / "data" / "catalog_export.xlsx"


def _load_portal_categories() -> dict:
    """Загружает items.json, возвращает by_model и by_name для подстановки категории."""
    if not PORTAL_ITEMS_JSON.exists():
        return {"by_model": {}, "by_name": {}}
    try:
        data = json.loads(PORTAL_ITEMS_JSON.read_text(encoding="utf-8"))
        items = data.get("items", [])
        by_model = {}
        by_name = {}
        for it in items:
            model = (it.get("model") or "").strip()
            name = (it.get("name") or "").strip()
            if model:
                by_model[model.upper()] = it
            if name:
                by_name[name.lower()] = it
        return {"by_model": by_model, "by_name": by_name}
    except Exception as e:
        logger.debug("Не удалось загрузить portal items.json: %s", e)
        return {"by_model": {}, "by_name": {}}


def _normalize_price_rrc(p: dict) -> float:
    """Нормализация розничной цены (логика как в main.py)."""
    price_rrc_raw = p.get("price_rrc")
    price_client_raw = p.get("price_client")
    if price_rrc_raw is None or price_rrc_raw == "" or float(price_rrc_raw or 0) == 0:
        if price_client_raw not in (None, "") and float(price_client_raw or 0) > 0:
            price_rrc_raw = price_client_raw
        else:
            price_rrc_raw = p.get("price") or p.get("cost") or p.get("retail_price")
    if price_rrc_raw is None or price_rrc_raw == "":
        price_rrc_raw = p.get("price") or p.get("cost") or p.get("retail_price") or 0
    if isinstance(price_rrc_raw, str):
        try:
            return float(price_rrc_raw.replace(",", "."))
        except (ValueError, AttributeError):
            return 0.0
    try:
        return float(price_rrc_raw)
    except (ValueError, TypeError):
        return 0.0


def run(output_path: Path, use_cache_fallback: bool = True) -> None:
    """Загружает товары из API Complex и записывает их в Excel."""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment
    except ImportError:
        logger.error("Установите openpyxl: pip install openpyxl")
        sys.exit(1)

    client = B2BClient(api_url=B2B_API_URL, api_key=API_KEY)
    data = client.update_products(use_cache=use_cache_fallback)
    products = data.get("products") or []
    if not products:
        logger.warning("Нет товаров. Проверьте API_KEY и доступность API или кэш.")
        sys.exit(1)

    logger.info("Загружено товаров из API/кэша: %s", len(products))
    portal_map = _load_portal_categories()
    by_model = portal_map.get("by_model", {})
    by_name = portal_map.get("by_name", {})

    # Подготовка строк для Excel: те же цены, что в каталоге
    rows = []
    for p in products:
        price_rrc = _normalize_price_rrc(p)
        model = (p.get("model") or "").strip()
        brand = (p.get("brand") or "").strip()
        name = (p.get("name") or "").strip()
        quantity = int(p.get("quantity") or 0)
        if quantity < 0:
            quantity = 0

        price_info = calculate_final_price(price_rrc, model, brand)
        final_price = float(price_info["final_price"])
        discount = float(price_info["discount"])
        discount_amount = float(price_info["discount_amount"])

        category = ""
        model_key = model.upper()
        name_key = name.lower()
        item = by_model.get(model_key) or by_name.get(name_key)
        if item:
            category = (item.get("category") or "").strip()

        rows.append({
            "name": name,
            "model": model,
            "brand": brand,
            "quantity": quantity,
            "price_rrc": price_rrc,
            "final_price": final_price,
            "discount": discount,
            "discount_amount": discount_amount,
            "category": category,
        })

    # Запись в Excel
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Каталог"

    headers = [
        "Наименование",
        "Модель",
        "Бренд",
        "Остаток",
        "Цена РРЦ",
        "Итоговая цена",
        "Скидка %",
        "Сумма скидки",
        "Категория",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for row_idx, r in enumerate(rows, start=2):
        ws.cell(row=row_idx, column=1, value=r["name"])
        ws.cell(row=row_idx, column=2, value=r["model"])
        ws.cell(row=row_idx, column=3, value=r["brand"])
        ws.cell(row=row_idx, column=4, value=r["quantity"])
        ws.cell(row=row_idx, column=5, value=r["price_rrc"])
        ws.cell(row=row_idx, column=6, value=r["final_price"])
        ws.cell(row=row_idx, column=7, value=r["discount"])
        ws.cell(row=row_idx, column=8, value=r["discount_amount"])
        ws.cell(row=row_idx, column=9, value=r["category"])

    # Ширина столбцов
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 35

    wb.save(output_path)
    logger.info("Сохранено: %s (строк: %s)", output_path, len(rows))


def main() -> None:
    import argparse
    parser = argparse.ArgumentParser(description="Экспорт каталога API Complex в Excel")
    parser.add_argument(
        "--output",
        "-o",
        type=Path,
        default=DEFAULT_OUTPUT,
        help="Путь к выходному .xlsx файлу",
    )
    parser.add_argument(
        "--no-cache",
        action="store_true",
        help="Не использовать кэш при недоступности API",
    )
    args = parser.parse_args()
    run(output_path=args.output, use_cache_fallback=not args.no_cache)


if __name__ == "__main__":
    main()
