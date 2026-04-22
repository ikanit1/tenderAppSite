#!/usr/bin/env python3
"""
Скрипт формирует Excel для импорта в SATU в формате XLS/XLSX по документации портала.

Важно: SATU не читает картинки из ячеек Excel. Импорт фото возможен только по URL в колонке
«Ссылка_изображения» (формат: https://..., несколько ссылок через запятую с пробелом).

Как передать фото в SATU:
  1) --image-via-api — ссылки на фото через ваш же API (main.py): /api/products/{модель}/image.
     Подходит, если apisite доступен по SITEMAP_BASE_URL (напр. https://grgroup.kz). Ничего выкладывать не нужно.
  2) --image-base-url URL — если выложите папки portal_export на другой сервер; в Excel подставятся URL/папка/файл.jpg
  3) После импорта Excel: python update_satu_products.py — фото дозагрузятся в SATU по API

- Вкладка «Export Products Sheet» — товары с обязательными полями SATU (формат по документации).
- Вкладка «Export Groups Sheet» — минимальная структура групп.

По умолчанию данные берутся из API (B2B) с подстановкой описания и изображений из portal_export.
  --from-api (по умолчанию) — загрузка из B2B + слияние с portal_export, цены со скидками.
  --no-from-api — только portal_export (как раньше).

Запуск:
  python export_satu_excel.py --image-via-api
  python export_satu_excel.py -o data/satu_import.xlsx --image-base-url "https://site.com/catalog/"
  python export_satu_excel.py --limit 100
  python export_satu_excel.py --no-from-api   # только из portal_export
"""
import json
import re
import sys
import time
from pathlib import Path
from urllib.parse import quote

_apisite_dir = Path(__file__).resolve().parent
if str(_apisite_dir) not in sys.path:
    sys.path.insert(0, str(_apisite_dir))

from upload_to_satu import load_products_from_portal, _price_to_tenge
from utils import model_to_foldername, get_clean_id, normalize_model_for_fs
from satu_categories import classify_product, get_all_group_names, get_group_number

try:
    import openpyxl
except ImportError:
    print("Установите openpyxl: pip install openpyxl")
    sys.exit(1)

# Ограничения SATU
MAX_NAZVA = 100
MAX_OPISANIE = 12160
MAX_POISK_ZAPROS = 255
MAX_KOD_TOVARA = 25
MAX_PROIZVODITEL = 255
MAX_IDENTIFIKATOR = 255

# Характеристики: blacklist ключей (не являются характеристиками товара)
_ATTRS_BLACKLIST = {"Итого", "Код Elevel"}

# Whitelist единиц измерения для парсинга значений характеристик
# Порядок важен: сначала длинные (мм² перед мм, кВт перед Вт)
_UNIT_PATTERN = re.compile(
    r"\s+(мм²|мм|см|м|км|мкм|г|кг|т|шт|кВт|МВт|Вт|кГц|МГц|Гц|°C|°|кОм|МОм|Ом|мкФ|нФ|пФ|Ф|Гн|лм|лк|кПа|МПа|Па|бар|дБ|мл|л|А|В)$"
)

MAX_SEO_TITLE = 250
MAX_SEO_DESC = 250


def _parse_attr_value(raw_value: str) -> tuple[str, str]:
    """Разделяет значение характеристики на (значение, единица_измерения).

    Примеры:
        '44.5 мм' → ('44.5', 'мм')
        'Да' → ('Да', '')
        '4 А' → ('4', 'А')
        '1.0...25.0 мм²' → ('1.0...25.0', 'мм²')
        '-20…+55' → ('-20…+55', '')
    """
    raw_value = raw_value.strip()
    if not raw_value:
        return ("", "")
    m = _UNIT_PATTERN.search(raw_value)
    if m:
        unit = m.group(1)
        value = raw_value[:m.start()].strip()
        return (value, unit)
    return (raw_value, "")


def _filter_attributes(attrs: dict) -> list[tuple[str, str, str]]:
    """Фильтрует и парсит attributes dict в список (название, измерение, значение).

    Убирает ключи из blacklist, парсит единицы измерения.
    """
    result = []
    for key, value in attrs.items():
        if key in _ATTRS_BLACKLIST:
            continue
        val_str = str(value).strip() if value else ""
        if not val_str:
            continue
        parsed_value, unit = _parse_attr_value(val_str)
        result.append((key, unit, parsed_value))
    return result


def _strip_html_tags(html_str: str) -> str:
    """Удаляет все HTML-теги из строки."""
    return re.sub(r"<[^>]+>", " ", html_str).strip()


def _build_seo_title(name: str, max_len: int = MAX_SEO_TITLE) -> str:
    """Генерирует HTML_заголовок для SEO.

    Шаблон: '{Название} — Купить в Астане по выгодной цене | G&R Group'
    """
    suffix = " — Купить в Астане по выгодной цене | G&R Group"
    if len(name) + len(suffix) <= max_len:
        return name + suffix
    available = max_len - len(suffix)
    truncated = name[:available].rsplit(" ", 1)[0] if available > 10 else name[:available]
    return truncated + suffix


def _build_seo_description(description: str, max_len: int = MAX_SEO_DESC) -> str:
    """Генерирует HTML_описание для SEO.

    Шаблон: '{первые ~150 символов без HTML}. Купить в Астане с доставкой. G&R Group'
    """
    suffix = ". Купить в Астане с доставкой. G&R Group"
    plain = _strip_html_tags(description)
    plain = re.sub(r"\s+", " ", plain).strip()
    available = max_len - len(suffix)
    if available <= 0:
        return suffix.lstrip(". ")
    if len(plain) > available:
        plain = plain[:available].rsplit(" ", 1)[0]
    if not plain:
        return "Купить в Астане с доставкой. G&R Group"
    return plain + suffix


_COUNTRY_KEYS = ("Страна", "Страна производства", "Страна-производитель", "Страна производитель", "Country")


def _extract_country(attrs: dict) -> str:
    """Извлекает страну производителя из attributes."""
    for key in _COUNTRY_KEYS:
        if key in attrs:
            return str(attrs[key]).strip()
    lower_map = {k.lower(): v for k, v in attrs.items()}
    for key in _COUNTRY_KEYS:
        if key.lower() in lower_map:
            return str(lower_map[key.lower()]).strip()
    return ""


_SCRAPE_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
}

_IMG_ACCEPT = {
    "User-Agent": _SCRAPE_HEADERS["User-Agent"],
    "Accept": "image/webp,image/png,image/jpeg,*/*;q=0.8",
}


def _get_original_image_url(cache_url: str) -> str:
    """Преобразует URL кеша complex.com.kz в оригинал.

    '/image/cache/catalog/.../main-360x360.jpg' → '/image/catalog/.../main.jpg'
    """
    url = cache_url.replace("/cache/", "/")
    url = re.sub(r"-\d+x\d+(\.\w+)$", r"\1", url)
    return url


def ensure_product_images(portal_export_dir: Path, max_products: int | None = None) -> dict[str, int]:
    """Скачивает фото с complex.com.kz для товаров без изображений.

    Возвращает статистику: {'checked': N, 'downloaded': N, 'failed': N, 'skipped': N}
    """
    from urllib.request import Request, urlopen
    from urllib.error import URLError, HTTPError

    stats = {"checked": 0, "downloaded": 0, "failed": 0, "skipped": 0}
    if not portal_export_dir.is_dir():
        return stats

    folders = sorted(f for f in portal_export_dir.iterdir() if f.is_dir())
    if max_products:
        folders = folders[:max_products]

    for folder in folders:
        product_json = folder / "product.json"
        if not product_json.exists():
            continue

        # Проверяем наличие image-файлов
        existing_images = [
            f for f in folder.iterdir()
            if f.is_file() and f.name.startswith("image") and f.suffix.lower() in (".jpg", ".jpeg", ".png", ".webp")
        ]
        if existing_images:
            stats["skipped"] += 1
            continue

        stats["checked"] += 1

        try:
            data = json.loads(product_json.read_bytes())
        except Exception:
            stats["failed"] += 1
            continue

        product_url = (data.get("product_url") or "").strip()
        if not product_url:
            stats["failed"] += 1
            continue

        try:
            req = Request(product_url, headers=_SCRAPE_HEADERS)
            resp = urlopen(req, timeout=15)
            html = resp.read().decode("utf-8", errors="replace")
        except (URLError, HTTPError, OSError) as e:
            print(f"  [WARN] HTTP error for {folder.name}: {e}", file=sys.stderr)
            stats["failed"] += 1
            continue

        # Ищем изображения в секции oct-gallery (основное фото + галерея)
        image_urls = re.findall(
            r'oct-gallery[^>]*href=["\']([^"\']+)["\']', html
        )
        if not image_urls:
            # Фоллбэк: ищем img src в productImages секции
            idx = html.find("productImages")
            if idx >= 0:
                section = html[idx:idx + 5000]
                image_urls = re.findall(
                    r'src=["\'](https?://complex\.com\.kz/image/[^"\']+\.(?:jpg|jpeg|png|webp))["\']',
                    section, re.I
                )

        # Дедупликация и фильтрация
        seen = set()
        clean_urls = []
        for url in image_urls:
            if not url.startswith("http"):
                url = "https://complex.com.kz" + url
            if any(x in url.lower() for x in ("logo", "icon", "new-icon", "shipproduct", "distributor")):
                continue
            if not re.search(r"\.\w{3,4}$", url):
                continue
            orig = _get_original_image_url(url)
            if orig not in seen:
                seen.add(orig)
                clean_urls.append(orig)

        if not clean_urls:
            stats["failed"] += 1
            continue

        # Скачиваем (макс. 10 фото)
        downloaded_names = []
        for i, img_url in enumerate(clean_urls[:10], 1):
            ext = Path(img_url).suffix.lower() or ".jpg"
            if ext not in (".jpg", ".jpeg", ".png", ".webp"):
                ext = ".jpg"
            filename = f"image_{i:02d}{ext}"
            try:
                req = Request(img_url, headers=_IMG_ACCEPT)
                img_resp = urlopen(req, timeout=15)
                img_data = img_resp.read()
                if len(img_data) < 500:
                    continue
                (folder / filename).write_bytes(img_data)
                downloaded_names.append(filename)
            except (URLError, HTTPError, OSError) as e:
                print(f"  [WARN] Failed to download image {img_url}: {e}", file=sys.stderr)
                continue

        if downloaded_names:
            data["images"] = downloaded_names
            data["images_count"] = len(downloaded_names)
            product_json.write_bytes(json.dumps(data, ensure_ascii=False, indent=2).encode("utf-8"))
            stats["downloaded"] += 1
            print(f"  [OK] {folder.name}: downloaded {len(downloaded_names)} images")
        else:
            stats["failed"] += 1

        time.sleep(0.5)

    return stats


# Заголовки вкладки «Export Products Sheet» (формат SATU)
HEADERS_PRODUCTS = (
    "Код_товара",
    "Название_позиции",
    "Поисковые_запросы",
    "Описание",
    "Тип_товара",
    "Цена",
    "Цена от",
    "Валюта",
    "Единица_измерения",
    "Минимальный_объем_заказа",
    "Оптовая_цена",
    "Минимальный_заказ_опт",
    "Количество",
    "Ссылка_изображения",
    "Наличие",
    "Идентификатор_товара",
    "Производитель",
    "Страна_производитель",
    "HTML_заголовок",
    "HTML_описание",
    "Номер_группы",
    "Название_группы",
)

PORTAL_EXPORT_DIR = _apisite_dir / "portal_export"


def _get_api_base_url() -> str:
    """Базовый URL сайта (apisite), откуда отдаются изображения по /api/products/{model}/image."""
    try:
        from config import SITEMAP_BASE_URL
        return (SITEMAP_BASE_URL or "").strip().rstrip("/") or ""
    except Exception:
        return ""


def _normalize_price_rrc(p: dict) -> float:
    """Нормализация розничной цены (логика как в main.py / export_catalog_to_excel)."""
    price_rrc_raw = p.get("price_rrc")
    price_client_raw = p.get("price_client")
    if price_rrc_raw is None or price_rrc_raw == "" or float(price_rrc_raw or 0) == 0:
        if price_client_raw not in (None, "") and float(price_client_raw or 0) > 0:
            price_rrc_raw = price_client_raw
        else:
            price_rrc_raw = p.get("price") or p.get("cost") or p.get("retail_price")
    if price_rrc_raw is None or price_rrc_raw == "":
        price_rrc_raw = p.get("price") or p.get("cost") or p.get("retail_price") or 0
    if isinstance(price_rrc_raw, str):
        try:
            return float(price_rrc_raw.replace(",", "."))
        except (ValueError, AttributeError):
            return 0.0
    try:
        return float(price_rrc_raw)
    except (ValueError, TypeError):
        return 0.0


def is_valid_price(price: float | None) -> bool:
    """Возвращает True только если цена реальная и адекватная. Иначе — пустая ячейка (цена по запросу)."""
    if price is None:
        return False
    if price <= 0.01:
        return False
    if price >= 9_000_000:
        return False
    return True


def _build_portal_by_model() -> dict:
    """
    Строит словарь для слияния с API: по разным ключам (model, folder_name, clean_id)
    возвращается {description, image_paths, folder} из portal_export.
    Обрабатывает оба формата: product.json (браузерный парсер) и info.json (старый парсер).
    """
    result = {}
    if not PORTAL_EXPORT_DIR.is_dir():
        return result

    def _scan_images(folder: Path) -> list:
        """Сканирует папку и возвращает список Path для image_* файлов."""
        return sorted(
            f for f in folder.iterdir()
            if f.suffix.lower() in (".jpg", ".jpeg", ".png", ".webp") and f.name.startswith("image")
        )

    def _register(registry: dict, model: str, folder_name: str, record: dict):
        """Регистрирует запись по нескольким ключам (model, folder_name, clean_id, foldername)."""
        for key in (model, folder_name, get_clean_id(normalize_model_for_fs(model)), model_to_foldername(model)):
            if key and key not in registry:
                registry[key] = record

    # --- Проход 1: папки с product.json (браузерный парсер) ---
    for folder in sorted(PORTAL_EXPORT_DIR.iterdir()):
        if not folder.is_dir():
            continue
        product_json = folder / "product.json"
        if not product_json.exists():
            continue
        try:
            data = json.loads(product_json.read_bytes())
        except Exception:
            continue

        name = (data.get("name") or "").strip()
        if not name and (folder / "name.txt").exists():
            name = (folder / "name.txt").read_bytes().decode("utf-8-sig", errors="replace").strip()
        if not name:
            name = folder.name
        model = (data.get("model") or "").strip() or folder.name
        brand = (data.get("brand") or data.get("manufacturer") or "").strip()

        desc_html = (data.get("description_html") or "").strip()
        desc_plain = (data.get("description") or "").strip()
        desc_file = folder / "description.txt"
        if desc_file.exists():
            desc_plain = desc_file.read_bytes().decode("utf-8-sig", errors="replace").strip() or desc_plain

        if desc_html:
            description = desc_html
        elif desc_plain:
            description = desc_plain
        else:
            attrs = data.get("attributes") or {}
            category = classify_product(model, name)
            description = _build_description_from_attrs(name, model, brand, category, attrs)

        description = description[:MAX_OPISANIE]

        image_names = data.get("images") or []
        image_paths = [folder / im for im in image_names if (folder / im).exists()]
        if not image_paths:
            image_paths = _scan_images(folder)

        attrs = data.get("attributes") or {}
        record = {"description": description, "image_paths": image_paths, "folder": folder.name, "attributes": attrs}
        _register(result, model, folder.name, record)

    # --- Проход 2: папки только с info.json (старый парсер без product.json) ---
    for folder in sorted(PORTAL_EXPORT_DIR.iterdir()):
        if not folder.is_dir():
            continue
        if (folder / "product.json").exists():
            continue  # уже обработано в проходе 1
        info_json = folder / "info.json"
        if not info_json.exists():
            continue
        try:
            data = json.loads(info_json.read_bytes())
        except Exception:
            continue

        model = (data.get("model") or "").strip() or folder.name
        name = (data.get("name") or "").strip() or model

        desc_file = folder / "description.txt"
        description = ""
        if desc_file.exists():
            description = desc_file.read_bytes().decode("utf-8-sig", errors="replace").strip()
        if not description:
            category = classify_product(model, name)
            description = _build_description_from_template(name, model, "", category)
        description = description[:MAX_OPISANIE]

        image_paths = _scan_images(folder)

        attrs = data.get("attributes") or {}
        record = {"description": description, "image_paths": image_paths, "folder": folder.name, "attributes": attrs}
        _register(result, model, folder.name, record)

    return result


def load_products_for_satu(from_api: bool = True, limit: int | None = None) -> list:
    """
    Возвращает список товаров для выгрузки в SATU.
    При from_api=True: загрузка из B2B API, нормализация цен, скидки (price_manager),
    слияние с portal_export по model (описание, image_paths). Поля: name, model, brand, quantity,
    final_price, description, image_paths, folder.
    При from_api=False: загрузка только из portal_export (как load_products_from_portal),
    с подстановкой final_price из price_tenge и brand из product.json или "".
    """
    if from_api:
        from b2b_client import B2BClient
        from price_manager import calculate_final_price

        client = B2BClient()
        data = client.update_products(use_cache=False)
        api_products = data.get("products") or []
        if not api_products:
            return []

        portal_by_model = _build_portal_by_model()

        def lookup_portal(model: str):
            if not model:
                return {}
            model = model.strip()
            for key in (model, model_to_foldername(model), get_clean_id(normalize_model_for_fs(model))):
                if key and key in portal_by_model:
                    return portal_by_model[key]
            return {}

        out = []
        for p in api_products:
            price_rrc = _normalize_price_rrc(p)
            model = (p.get("model") or "").strip()
            brand = (p.get("brand") or "").strip()
            name_raw = (p.get("name") or "").strip()
            if model and name_raw.startswith(model):
                name = name_raw[len(model):].strip()
            else:
                name = name_raw
            quantity = max(0, int(p.get("quantity") or 0))

            portal = lookup_portal(model)
            # Если price_rrc нулевая/некорректная — пробуем взять цену из portal_export/product.json
            if not is_valid_price(price_rrc):
                portal_folder = portal.get("folder") or model
                portal_json = PORTAL_EXPORT_DIR / portal_folder / "product.json"
                if portal_json.exists():
                    try:
                        portal_data = json.loads(portal_json.read_bytes())
                        portal_price = _price_to_tenge(portal_data.get("price") or 0)
                        if is_valid_price(portal_price):
                            price_rrc = portal_price
                    except Exception:
                        pass

            price_info = calculate_final_price(price_rrc, model, brand)
            final_price = float(price_info["final_price"])
            if not is_valid_price(final_price) and is_valid_price(price_rrc):
                final_price = price_rrc
            description = (portal.get("description") or "").strip()
            if not description:
                category = classify_product(model, name)
                description = _build_description_from_template(name, model, brand, category)
            image_paths = portal.get("image_paths", [])
            folder = portal.get("folder", "")

            out.append({
                "name": name,
                "model": model,
                "brand": brand,
                "quantity": quantity,
                "price_rrc": price_rrc,
                "final_price": final_price,
                "description": description,
                "image_paths": image_paths,
                "folder": folder,
                "attributes": portal.get("attributes", {}),
            })
            if limit is not None and len(out) >= limit:
                break
        return out

    # from_api=False: только portal_export
    portal_list = load_products_from_portal(limit=limit)
    from price_manager import calculate_final_price

    out = []
    for p in portal_list:
        model = (p.get("model") or "").strip()
        name = (p.get("name") or "").strip()
        brand = ""
        try:
            portal_dir = PORTAL_EXPORT_DIR / p.get("folder", "")
            if (portal_dir / "product.json").exists():
                data = json.loads((portal_dir / "product.json").read_bytes())
                brand = (data.get("brand") or data.get("manufacturer") or "").strip()
        except Exception:
            pass
        attrs = {}
        try:
            portal_dir = PORTAL_EXPORT_DIR / p.get("folder", "")
            if (portal_dir / "product.json").exists():
                pdata = json.loads((portal_dir / "product.json").read_bytes())
                attrs = pdata.get("attributes") or {}
        except Exception:
            pass
        price_tenge = float(p.get("price_tenge") or 0)
        price_info = calculate_final_price(price_tenge, model, brand)
        final_price = float(price_info["final_price"])

        out.append({
            "name": name,
            "model": model,
            "brand": brand,
            "quantity": max(0, int(p.get("quantity", 1))),
            "price_rrc": price_tenge,
            "final_price": final_price,
            "description": p.get("description", ""),
            "image_paths": p.get("image_paths", []),
            "folder": p.get("folder", ""),
            "attributes": attrs,
        })
    return out


def _build_search_queries(name: str, brand: str = "", category: str = "", max_len: int = MAX_POISK_ZAPROS) -> str:
    """Формирует поисковые запросы: слова из названия + бренд + категория + фразы покупателя."""
    if not name:
        return "товар"
    parts = []
    words = re.findall(r"[а-яёa-z0-9\-]+", name.lower(), re.I)
    words = [w[:50] for w in words if len(w) >= 2][:10]
    parts.extend(words)
    if brand and brand.lower() not in name.lower():
        parts.append(brand.lower())
    if category:
        parts.append(category.lower())
        parts.append(f"{category.lower()} купить")
        if brand:
            parts.append(f"{brand.lower()} {category.lower()}")
    s = ", ".join(dict.fromkeys(parts))  # убираем дубли, сохраняем порядок
    return s[:max_len] if s else "товар"


# Шаблоны описаний по категориям SATU
_CATEGORY_DESCRIPTIONS: dict[str, str] = {
    "Автоматические выключатели": "модульный аппарат защиты для распределительных щитов. Предназначен для защиты электрических цепей от токов перегрузки и короткого замыкания",
    "Дифференциальная защита": "устройство дифференциальной защиты (УЗО/АВДТ) для защиты от токов утечки и поражения электрическим током",
    "Контакторы и пускатели": "электромагнитный контактор для коммутации силовых цепей и управления электродвигателями",
    "Видеокамеры": "камера видеонаблюдения для круглосуточной записи и мониторинга объектов в реальном времени",
    "Видеорегистраторы": "цифровой видеорегистратор для записи, хранения и воспроизведения видео с камер наблюдения",
    "Коммутаторы": "сетевой коммутатор для построения локальных сетей и высокоскоростной передачи данных",
    "Точки доступа и маршрутизаторы": "устройство беспроводного доступа для организации сети Wi-Fi и передачи данных",
    "Кабель и провод": "кабельная продукция для монтажа электрических сетей, линий связи и управления",
    "Кабельные каналы": "кабельный канал для прокладки и защиты проводов и кабелей в помещениях",
    "Лотки и аксессуары лотков": "кабельный лоток для организованной прокладки кабелей в производственных и офисных помещениях",
    "Светодиодные лампы": "светодиодная лампа для энергоэффективного освещения с длительным сроком службы",
    "Светильники и прожекторы": "осветительный прибор для освещения производственных, офисных и бытовых помещений",
    "Светодиодные ленты": "светодиодная лента для декоративной и функциональной подсветки помещений и объектов",
    "Наконечники и гильзы": "обжимной наконечник или гильза для надёжного подключения проводников к клеммам",
    "Розетки и выключатели": "электроустановочное изделие для подключения потребителей и управления освещением",
    "Датчики": "датчик для автоматического обнаружения и измерения физических параметров",
    "Реле": "электромеханическое реле для управления цепями по сигналу от датчиков и контроллеров",
    "Электродвигатели": "асинхронный электродвигатель для привода производственных механизмов и оборудования",
    "Шкафы и щиты": "монтажный корпус для размещения и защиты электрооборудования и щитовой аппаратуры",
    "Трансформаторы": "силовой трансформатор для преобразования напряжения в электрических цепях",
    "Разъемы и коннекторы": "соединительный разъём для надёжной стыковки кабелей и электрических цепей",
    "Блоки питания": "блок питания для обеспечения стабилизированного постоянного напряжения",
    "Зажимы и клеммы": "клеммное соединение для надёжного и безопасного подключения проводников",
    "Монтажные коробки": "распределительная коробка для соединения и защиты электрических проводников",
    "Частотные преобразователи": "преобразователь частоты для плавного пуска и бесступенчатого регулирования скорости электродвигателей",
    "Муфты и трубы": "защитная труба или муфта для прокладки кабелей в условиях механических воздействий",
    "Устройства плавного пуска": "устройство плавного пуска для снижения пускового тока и механических нагрузок на двигатель",
    "Кнопки управления": "кнопка или переключатель для дистанционного управления электрооборудованием",
    "Контроль доступа (СКУД)": "оборудование системы контроля и управления доступом для обеспечения безопасности объектов",
    "Аудио и видеотехника": "аудио- или видеооборудование для трансляции, записи и воспроизведения сигналов",
    "Патроны и стартёры": "патрон или стартёр для подключения и запуска источников освещения",
    "Счётчики электроэнергии": "счётчик электрической энергии для учёта потребления электроэнергии",
    "Жёсткие диски": "жёсткий диск или SSD-накопитель для хранения данных в системах видеонаблюдения и серверах",
    "Кронштейны и крепёж": "монтажный кронштейн или крепёжный элемент для установки оборудования",
    "Мониторы и дисплеи": "монитор или дисплей для отображения видео и управления системами наблюдения",
    "Прочее": "электротехническое оборудование и компоненты для промышленного и бытового применения",
}


def _build_description_from_template(name: str, model: str, brand: str, category: str) -> str:
    """Генерирует HTML-описание товара на основе категорийного шаблона.

    Используется как финальный фоллбэк когда description_html, description.txt
    и attributes все пустые. Соответствует требованиям SATU к полю «Описание».
    """
    category_text = _CATEGORY_DESCRIPTIONS.get(category) or _CATEGORY_DESCRIPTIONS["Прочее"]
    items = []
    if brand:
        items.append(f"<li><b>Производитель:</b> {brand}</li>")
    if model:
        items.append(f"<li><b>Артикул:</b> {model}</li>")
    ul = f"<ul>{''.join(items)}</ul>" if items else ""
    return f"<p><b>{name}</b> — {category_text}.</p>{ul}"


def _build_description_from_attrs(
    name: str, model: str, brand: str, category: str, attrs: dict
) -> str:
    """Генерирует HTML-описание из словаря атрибутов товара.

    Используется когда нет description_html и description.txt,
    но есть атрибуты. Строит: заголовок + категорийный абзац + HTML-таблица.
    Атрибуты из _ATTRS_BLACKLIST исключаются.
    """
    if not attrs:
        return _build_description_from_template(name, model, brand, category)

    category_text = _CATEGORY_DESCRIPTIONS.get(category) or _CATEGORY_DESCRIPTIONS["Прочее"]

    rows = []
    if brand:
        rows.append(f"<tr><th>Производитель</th><td>{brand}</td></tr>")
    if model:
        rows.append(f"<tr><th>Артикул</th><td>{model}</td></tr>")
    for key, value in attrs.items():
        if key in _ATTRS_BLACKLIST:
            continue
        val_str = str(value).strip() if value else ""
        if not val_str:
            continue
        rows.append(f"<tr><th>{key}</th><td>{val_str}</td></tr>")

    if not rows:
        return _build_description_from_template(name, model, brand, category)

    table = "<table>" + "".join(rows) + "</table>"
    result = f"<h3>{name}</h3><p>{category_text}.</p>{table}"
    return result


def _image_urls_for_product(product: dict, base_url: str) -> str:
    """Формирует строку URL изображений для SATU: «url1, url2». base_url — с завершающим слэшем (раздача по путям папка/файл)."""
    image_paths = product.get("image_paths") or []
    folder = (product.get("folder") or "").strip()
    if not base_url or not image_paths:
        return ""
    base = base_url.rstrip("/") + "/"
    parts = []
    for path in image_paths:
        name = path.name if hasattr(path, "name") else Path(path).name
        url = base + quote(folder) + "/" + quote(name)
        parts.append(url)
    return ", ".join(parts)


def _image_urls_via_api(product: dict, api_base_url: str) -> str:
    """Ссылки на изображения: https://grgroup.kz/catalog/api/products/{model}/image?index=0, ?index=1, ...
    api_base_url — базовый URL (напр. https://grgroup.kz)."""
    image_paths = product.get("image_paths") or []
    model = (product.get("model") or "").strip()
    if not api_base_url or not model or not image_paths:
        return ""
    base = api_base_url.rstrip("/")
    parts = []
    for i in range(len(image_paths)):
        url = f"{base}/catalog/api/products/{quote(model)}/image?index={i}"
        parts.append(url)
    return ", ".join(parts)


def build_full_excel(
    products: list,
    limit: int | None = None,
    image_base_url: str | None = None,
    image_via_api: bool = False,
    api_base_url: str | None = None,
):
    """Собирает Excel в формате SATU: Export Products Sheet + Export Groups Sheet.
    Ожидает в каждом товаре: name, model, brand, quantity, final_price, description, image_paths, folder.
    image_base_url: если задан, в Ссылка_изображения подставляются URL вида base_url/папка/файл.jpg
    image_via_api: если True, ссылки строятся на эндпоинт /api/products/{model}/image."""
    if limit:
        products = products[:limit]
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # --- Подготовка характеристик: определяем макс. кол-во блоков ---
    max_attrs = 0
    for p in products:
        attrs = p.get("attributes") or {}
        filtered = [k for k in attrs if k not in _ATTRS_BLACKLIST and str(attrs[k]).strip()]
        if len(filtered) > max_attrs:
            max_attrs = len(filtered)

    # --- Вкладка «Export Products Sheet» ---
    ws_products = wb.create_sheet("Export Products Sheet", 0)
    # Основные заголовки
    for col, h in enumerate(HEADERS_PRODUCTS, 1):
        c = ws_products.cell(row=1, column=col, value=h)
        c.font = openpyxl.styles.Font(bold=True)
    # Заголовки характеристик (блоки по 3 столбца)
    base_col_header = len(HEADERS_PRODUCTS) + 1
    for i in range(max_attrs):
        for j, suffix in enumerate(("Название_Характеристики", "Измерение_Характеристики", "Значение_Характеристики")):
            col_idx = base_col_header + i * 3 + j
            c = ws_products.cell(row=1, column=col_idx, value=suffix)
            c.font = openpyxl.styles.Font(bold=True)
    ws_products.row_dimensions[1].height = 20

    api_base = api_base_url or _get_api_base_url()
    col_w = [14, 40, 18, 50, 12, 12, 8, 10, 8, 10, 14, 12, 8, 50, 10, 28, 28, 20, 40, 40, 12, 30]

    seen_identifiers = {}  # Для дедупликации идентификаторов товара

    for row_idx, p in enumerate(products, 2):
        name = (p.get("name") or "")[:MAX_NAZVA]
        model = (p.get("model") or "")[:MAX_KOD_TOVARA]
        desc = (p.get("description") or "")[:MAX_OPISANIE]
        price = float(p.get("final_price") or 0)
        price_rrc = float(p.get("price_rrc") or 0)
        qty = max(0, int(p.get("quantity", 1)))
        brand = (p.get("brand") or "")[:MAX_PROIZVODITEL]
        category = classify_product(model, name)
        search = _build_search_queries(name, brand, category)
        attrs_raw = p.get("attributes") or {}

        if image_base_url:
            link_izobr = _image_urls_for_product(p, image_base_url)
        elif api_base and (p.get("image_paths") or []):
            link_izobr = _image_urls_via_api(p, api_base)
        else:
            link_izobr = ""

        # Идентификатор_товара: дедупликация для вариантов одной модели (MVA31-* и др.)
        base_ident = (p.get("model") or "")[:MAX_IDENTIFIKATOR]
        if base_ident in seen_identifiers:
            seen_identifiers[base_ident] += 1
            ident = f"{base_ident}_{seen_identifiers[base_ident]}"[:MAX_IDENTIFIKATOR]
        else:
            seen_identifiers[base_ident] = 1
            ident = base_ident

        # Цена: розничная (final_price со скидкой)
        price_value = price if is_valid_price(price) else None

        # Оптовая_цена: должна быть ≤ розничной (SATU валидация)
        if is_valid_price(price):
            opt_price_val = round(price * 0.85, 2)
            if opt_price_val > price:
                opt_price_val = price
        else:
            opt_price_val = None

        min_order_unique = 1
        min_order_opt_val = 2

        # SEO и страна
        country = _extract_country(attrs_raw)
        seo_title = _build_seo_title(name)
        seo_desc = _build_seo_description(desc)

        # Основные столбцы (22)
        ws_products.cell(row=row_idx, column=1, value=model)
        ws_products.cell(row=row_idx, column=2, value=name)
        ws_products.cell(row=row_idx, column=3, value=search)
        ws_products.cell(row=row_idx, column=4, value=desc)
        ws_products.cell(row=row_idx, column=5, value="u")
        ws_products.cell(row=row_idx, column=6, value=price_value)
        ws_products.cell(row=row_idx, column=7, value="-")
        ws_products.cell(row=row_idx, column=8, value="KZT")
        ws_products.cell(row=row_idx, column=9, value="шт.")
        ws_products.cell(row=row_idx, column=10, value=min_order_unique)
        ws_products.cell(row=row_idx, column=11, value=opt_price_val)
        ws_products.cell(row=row_idx, column=12, value=min_order_opt_val)
        ws_products.cell(row=row_idx, column=13, value=qty)
        ws_products.cell(row=row_idx, column=14, value=link_izobr)
        ws_products.cell(row=row_idx, column=15, value="+" if qty > 0 else "-")
        ws_products.cell(row=row_idx, column=16, value=ident)
        ws_products.cell(row=row_idx, column=17, value=brand)
        ws_products.cell(row=row_idx, column=18, value=country)
        ws_products.cell(row=row_idx, column=19, value=seo_title)
        ws_products.cell(row=row_idx, column=20, value=seo_desc)
        ws_products.cell(row=row_idx, column=21, value=get_group_number(category))
        ws_products.cell(row=row_idx, column=22, value=category)

        # Характеристики (динамические столбцы)
        parsed_attrs = _filter_attributes(attrs_raw)
        base_col = len(HEADERS_PRODUCTS) + 1
        for attr_idx, (attr_name, attr_unit, attr_value) in enumerate(parsed_attrs):
            col_offset = base_col + attr_idx * 3
            ws_products.cell(row=row_idx, column=col_offset, value=attr_name)
            ws_products.cell(row=row_idx, column=col_offset + 1, value=attr_unit)
            ws_products.cell(row=row_idx, column=col_offset + 2, value=attr_value)

    for col, w in enumerate(col_w, 1):
        ws_products.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w
    # Ширина столбцов характеристик
    base_col = len(HEADERS_PRODUCTS) + 1
    for i in range(max_attrs):
        ws_products.column_dimensions[openpyxl.utils.get_column_letter(base_col + i * 3)].width = 30
        ws_products.column_dimensions[openpyxl.utils.get_column_letter(base_col + i * 3 + 1)].width = 12
        ws_products.column_dimensions[openpyxl.utils.get_column_letter(base_col + i * 3 + 2)].width = 25

    # --- Вкладка «Export Groups Sheet» ---
    ws_groups = wb.create_sheet("Export Groups Sheet", 1)
    headers_groups = ["Номер_группы", "Название_группы", "Идентификатор_группы", "Номер_родителя", "Идентификатор_родителя"]
    for col, h in enumerate(headers_groups, 1):
        c = ws_groups.cell(row=1, column=col, value=h)
        c.font = openpyxl.styles.Font(bold=True)
    for grp_idx, grp_name in enumerate(get_all_group_names(), 1):
        ws_groups.cell(row=grp_idx + 1, column=1, value=grp_idx)
        ws_groups.cell(row=grp_idx + 1, column=2, value=grp_name)
        for col in range(3, 6):
            ws_groups.cell(row=grp_idx + 1, column=col, value="")
    for col in range(1, 6):
        ws_groups.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 25

    return wb, len(products)


def main():
    import argparse
    parser = argparse.ArgumentParser(description="Собрать Excel в формате SATU (данные из API или portal_export)")
    parser.add_argument("-o", "--output", type=Path, default=_apisite_dir / "satu_import_full.xlsx", help="Путь к .xlsx")
    parser.add_argument("--limit", type=int, default=None, help="Максимум товаров")
    parser.add_argument(
        "--from-api",
        action="store_true",
        default=True,
        help="Брать товары из B2B API и слияние с portal_export (по умолчанию).",
    )
    parser.add_argument(
        "--no-from-api",
        action="store_false",
        dest="from_api",
        help="Брать товары только из portal_export.",
    )
    parser.add_argument(
        "--image-via-api",
        action="store_true",
        help="Ссылки на фото через API apisite (GET /api/products/модель/image). Используется SITEMAP_BASE_URL из .env (напр. https://grgroup.kz).",
    )
    parser.add_argument(
        "--api-base-url",
        type=str,
        default=None,
        help="Базовый URL API (если не задан, берётся SITEMAP_BASE_URL из config). Нужен для --image-via-api при другом домене.",
    )
    parser.add_argument(
        "--image-base-url",
        type=str,
        default=None,
        help="Базовый URL раздачи по путям папка/файл (напр. https://site.com/catalog/). Загрузите папки portal_export по этому адресу.",
    )
    parser.add_argument(
        "--skip-images",
        action="store_true",
        help="Пропустить загрузку недостающих фото с complex.com.kz.",
    )
    args = parser.parse_args()

    if not args.skip_images:
        print("Проверяем наличие фото товаров...")
        img_stats = ensure_product_images(PORTAL_EXPORT_DIR)
        print(f"Фото: проверено={img_stats['checked']}, загружено={img_stats['downloaded']}, "
              f"пропущено={img_stats['skipped']}, ошибок={img_stats['failed']}")

    products = load_products_for_satu(from_api=args.from_api, limit=args.limit)
    if not products:
        msg = "Товары не найдены в API/кэше." if args.from_api else "Товары не найдены в portal_export."
        print(msg)
        sys.exit(1)
    wb, count = build_full_excel(
        products,
        limit=args.limit,
        image_base_url=args.image_base_url,
        image_via_api=args.image_via_api,
        api_base_url=args.api_base_url,
    )
    args.output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.output)
    print(f"Готово: {count} позиций сохранено в {args.output}")


if __name__ == "__main__":
    main()
